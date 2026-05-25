"""
==============================================================
  Postilion Journal Reconciliation Tool
  excel_check.py
==============================================================
  Reconciles generated PDFs against the Excel transaction
  ledger exported from the core banking system.

  What it does:
  1. Scans output_folder for all generated PDF files
  2. Reads all .xlsx files in ExcelFiles/ folder
  3. Matches PDFs to Excel rows using BIN + RRN prefix
  4. Highlights matched rows in yellow in the Excel file
  5. Deletes any PDFs that have no matching Excel record
  6. Removes empty BIN subfolders from output
  7. Writes full audit trail to logs/

  Expected Excel columns:
    PAN                    — masked card number (e.g. 412345******7890)
    RETRIEVAL_REFERENCE_NR — transaction RRN (10 or 12 digits)

  Folder structure expected:
    ExcelFiles/     <- place reconciliation .xlsx files here
    output_folder/  <- PDFs from journal_split.py
    logs/           <- log file written here automatically

  Usage:
    python excel_check.py
    OR double-click run_excel_check.bat (Windows)
    Run AFTER journal_split.py has completed.
==============================================================
"""

import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#  FOLDER PATHS 

EXCEL_FOLDER  = "ExcelFiles"
OUTPUT_FOLDER = "output_folder"
LOGS_FOLDER   = "logs"

#  LOGGING SETUP 

os.makedirs(LOGS_FOLDER, exist_ok=True)

run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_filename  = os.path.join(LOGS_FOLDER, f"excel_check_{run_timestamp}.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_filename, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

#  HIGHLIGHT COLOUR

MATCH_COLOUR = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)

#  COUNTERS 

stats = {
    "excel_files_processed": 0,
    "excel_rows_checked":    0,
    "rows_matched":          0,
    "rows_unmatched":        0,
    "pdfs_total":            0,
    "pdfs_matched":          0,
    "pdfs_deleted":          0,
    "folders_removed":       0,
    "errors":                0,
}

#  UTILITIES 

def extract_bin(pan):
    """
    Extract first 6 digits (BIN) from a masked PAN string.
    Handles formats like: 412345******7890 or 412345XXXXXX7890
    """
    if pan and str(pan) not in ('nan', 'None', ''):
        return str(pan).split('*')[0][:6]
    return ""


def find_all_pdfs(output_folder):
    """Recursively find all PDF files in output_folder."""
    pdfs = []
    for root, dirs, files in os.walk(output_folder):
        for f in files:
            if f.endswith(".pdf"):
                pdfs.append(os.path.join(root, f))
    return pdfs


#  CORE RECONCILIATION 

def reconcile_excel_against_pdfs(excel_path, pdf_files):
    """
    For each row in the Excel ledger, extract BIN and RRN,
    check whether a matching PDF exists, and highlight if found.

    Match logic:
      PDF filename format : BIN_RRN_Date_Time.pdf
      Match key           : BIN_RRN (dynamic length — handles
                            both 10-digit and 12-digit RRNs)
    """
    matched_keys = []
    log.info(f"  Processing: {excel_path}")

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        stats["errors"] += 1
        log.error(f"  Cannot open {excel_path}: {e}")
        return matched_keys

    for sheet_name in wb.sheetnames:
        ws     = wb[sheet_name]
        header = [str(cell.value) for cell in ws[1]]

        if 'PAN' not in header or 'RETRIEVAL_REFERENCE_NR' not in header:
            log.warning(f"  Sheet '{sheet_name}': columns PAN / "
                        f"RETRIEVAL_REFERENCE_NR not found — skipping.")
            continue

        pan_col = header.index('PAN') + 1
        rrn_col = header.index('RETRIEVAL_REFERENCE_NR') + 1
        log.info(f"  Sheet: '{sheet_name}' | Rows: {ws.max_row - 1}")

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            stats["excel_rows_checked"] += 1

            pan_val = row[pan_col - 1]
            rrn_val = row[rrn_col - 1]

            if not pan_val or not rrn_val:
                continue

            pan = str(pan_val).strip()
            rrn = str(int(rrn_val)) if isinstance(rrn_val, float) else str(rrn_val).strip()

            bin_prefix = extract_bin(pan)
            if not bin_prefix or not rrn or rrn == '0':
                continue

            # Build match key — BIN_RRN (dynamic length)
            match_key  = f"{bin_prefix}_{rrn}"
            key_length = len(match_key)

            # Compare against start of each PDF filename
            matching_pdf = None
            for pdf_path in pdf_files:
                pdf_prefix = os.path.basename(pdf_path)[:key_length]
                if pdf_prefix == match_key:
                    matching_pdf = pdf_path
                    break

            if matching_pdf:
                for cell in ws[row_idx]:
                    cell.fill = MATCH_COLOUR
                matched_keys.append(match_key)
                stats["rows_matched"] += 1
                log.debug(f"  Row {row_idx}: MATCHED — {match_key}")
            else:
                stats["rows_unmatched"] += 1
                log.debug(f"  Row {row_idx}: no match — {match_key}")

    try:
        wb.save(excel_path)
        stats["excel_files_processed"] += 1
        log.info(f"  Saved with highlights: {excel_path}")
    except Exception as e:
        stats["errors"] += 1
        log.error(f"  Could not save {excel_path}: {e}")

    return matched_keys


# CLEANUP

def delete_unmatched_pdfs(output_folder, matched_keys):
    """Delete PDFs whose BIN_RRN prefix was not found in any Excel file."""
    pdf_files          = find_all_pdfs(output_folder)
    stats["pdfs_total"] = len(pdf_files)
    log.info(f"Total PDFs in output: {len(pdf_files)}")
    log.info(f"Total matched keys  : {len(matched_keys)}")

    for pdf_path in pdf_files:
        kept = False
        for key in matched_keys:
            if os.path.basename(pdf_path).startswith(key):
                stats["pdfs_matched"] += 1
                kept = True
                break
        if not kept:
            try:
                os.remove(pdf_path)
                stats["pdfs_deleted"] += 1
                log.info(f"  Deleted (unmatched): {os.path.basename(pdf_path)}")
            except Exception as e:
                stats["errors"] += 1
                log.error(f"  Could not delete {pdf_path}: {e}")


def remove_empty_folders(output_folder):
    """Remove BIN subfolders that are empty after PDF deletion."""
    for folder in os.listdir(output_folder):
        folder_path = os.path.join(output_folder, folder)
        if os.path.isdir(folder_path) and not os.listdir(folder_path):
            try:
                os.rmdir(folder_path)
                stats["folders_removed"] += 1
                log.info(f"  Removed empty folder: {folder_path}")
            except Exception as e:
                stats["errors"] += 1
                log.error(f"  Could not remove {folder_path}: {e}")


# MAIN

if __name__ == "__main__":
    start_time = datetime.now()

    log.info("=" * 60)
    log.info("  POSTILION JOURNAL RECONCILIATION — STARTED")
    log.info(f"  Run ID    : {run_timestamp}")
    log.info(f"  Excel dir : {os.path.abspath(EXCEL_FOLDER)}")
    log.info(f"  PDF dir   : {os.path.abspath(OUTPUT_FOLDER)}")
    log.info(f"  Log file  : {os.path.abspath(log_filename)}")
    log.info("=" * 60)

    # Step 1 — gather all PDFs
    all_pdfs = find_all_pdfs(OUTPUT_FOLDER)
    log.info(f"PDFs found: {len(all_pdfs)}")

    # Step 2 — reconcile each Excel file
    log.info("")
    log.info("=" * 60)
    log.info("STAGE 1 — EXCEL MATCHING AND ROW HIGHLIGHTING")
    log.info("=" * 60)

    all_matched_keys = []
    if not os.path.exists(EXCEL_FOLDER):
        log.error(f"Excel folder not found: {EXCEL_FOLDER}")
    else:
        excel_files = [f for f in os.listdir(EXCEL_FOLDER) if f.endswith(".xlsx")]
        log.info(f"Excel files found: {len(excel_files)}")
        for file in excel_files:
            matched = reconcile_excel_against_pdfs(
                os.path.join(EXCEL_FOLDER, file), all_pdfs
            )
            all_matched_keys.extend(matched)
            log.info(f"  Matches from {file}: {len(matched)}")

    # Step 3 — delete unmatched PDFs
    log.info("")
    log.info("=" * 60)
    log.info("STAGE 2 — DELETE UNMATCHED PDFs")
    log.info("=" * 60)
    delete_unmatched_pdfs(OUTPUT_FOLDER, all_matched_keys)

    # Step 4 — remove empty folders
    log.info("")
    log.info("=" * 60)
    log.info("STAGE 3 — CLEAN UP EMPTY FOLDERS")
    log.info("=" * 60)
    remove_empty_folders(OUTPUT_FOLDER)

    elapsed = (datetime.now() - start_time).seconds

    log.info("")
    log.info("=" * 60)
    log.info("  RECONCILIATION SUMMARY")
    log.info("=" * 60)
    log.info(f"  Excel files processed  : {stats['excel_files_processed']}")
    log.info(f"  Excel rows checked     : {stats['excel_rows_checked']}")
    log.info(f"  Rows matched           : {stats['rows_matched']}")
    log.info(f"  Rows unmatched         : {stats['rows_unmatched']}")
    log.info(f"  PDFs total             : {stats['pdfs_total']}")
    log.info(f"  PDFs matched (kept)    : {stats['pdfs_matched']}")
    log.info(f"  PDFs deleted           : {stats['pdfs_deleted']}")
    log.info(f"  Empty folders removed  : {stats['folders_removed']}")
    log.info(f"  Errors                 : {stats['errors']}")
    log.info(f"  Runtime                : {elapsed} seconds")
    log.info(f"  Log saved to           : {log_filename}")
    log.info("=" * 60)
    log.info("  RECONCILIATION COMPLETE")
    log.info("=" * 60)
